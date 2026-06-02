"""
API Python para processar ICMS ST usando openpyxl (baseado em Retencao_Autonoma.py).

Recebe modelo Excel + XMLs via upload multipart, valida, processa e retorna o
arquivo .xlsx gerado preservando fórmulas e tabelas nomeadas.

Segurança aplicada:
    - Sem modo debug em produção (Werkzeug debugger PIN é vetor de RCE).
    - CORS restritivo (somente origins listadas via env).
    - Limite global de payload (MAX_CONTENT_LENGTH).
    - Parser lxml com resolve_entities=False (anti-XXE / billion laughs).
    - Limpeza garantida do diretório temporário no finally.
    - Token compartilhado opcional via header X-API-Key.
    - Logging estruturado (sem print) e nível configurável por env.
"""

from __future__ import annotations

import logging
import os
import re
import shutil
import tempfile
from collections import Counter
from datetime import datetime
from html import unescape
from pathlib import Path

from flask import Flask, Response, jsonify, request, send_file
from flask_cors import CORS
from lxml import etree
from openpyxl import load_workbook
from openpyxl.cell.cell import MergedCell
from openpyxl.utils.cell import (
    column_index_from_string,
    coordinate_from_string,
)
from werkzeug.exceptions import RequestEntityTooLarge
from werkzeug.utils import secure_filename


# ===================== Configuração =====================

LOG_LEVEL = os.environ.get("ICMS_API_LOG_LEVEL", "INFO").upper()
logging.basicConfig(
    level=LOG_LEVEL,
    format="%(asctime)s [%(levelname)s] %(name)s: %(message)s",
)
logger = logging.getLogger("icms_api")

# Origins permitidos (CSV via env). Default: localhost dev.
DEFAULT_CORS = "http://localhost:3000,http://127.0.0.1:3000,http://localhost:5500"
ALLOWED_ORIGINS = [
    o.strip() for o in os.environ.get("ICMS_API_CORS_ORIGINS", DEFAULT_CORS).split(",") if o.strip()
]

# Token opcional. Se setado, requests precisam enviar X-API-Key.
API_KEY = os.environ.get("ICMS_API_KEY", "").strip()

# Limites de payload (default 25 MB total).
MAX_UPLOAD_MB = int(os.environ.get("ICMS_API_MAX_UPLOAD_MB", "25"))

# Limites de processamento.
MAX_XML_FILES = int(os.environ.get("ICMS_API_MAX_XML_FILES", "500"))


# ===================== App =====================

app = Flask(__name__)
app.config["MAX_CONTENT_LENGTH"] = MAX_UPLOAD_MB * 1024 * 1024
CORS(
    app,
    resources={r"/api/*": {"origins": ALLOWED_ORIGINS}},
    methods=["GET", "POST", "OPTIONS"],
    allow_headers=["Content-Type", "X-API-Key"],
    max_age=3600,
)


# ===================== Constantes de domínio =====================

NS = {"nfe": "http://www.portalfiscal.inf.br/nfe"}
GRUPOS: dict[str, dict[str, set[str]]] = {
    "1,54%.txt": {"cst": {"20"}, "csosn": set()},
    "4%.txt": {"cst": {"00"}, "csosn": set()},
    "7%.txt": {"cst": set(), "csosn": {"101", "102"}},
}
UF_VALIDO = "23"
CFOP_VALIDOS = {"5101", "5102", "5103", "5105", "5910"}
MAPEAMENTO_ABAS_CELULAS = {
    "1,54%.txt": ("Aliquota 1,54%", "D2"),
    "4%.txt": ("Aliquota 4%", "D2"),
    "7%.txt": ("Aliquota 7%", "D2"),
}

TEMP_DIR = Path(tempfile.gettempdir()) / "icms_api"

# Biblioteca de razões sociais (substituição manual via arquivo texto).
# Formato por linha: "NOME ORIGINAL => Nome Substituto" (igual Retencao_Autonoma.py).
# Path configurável via env. Default: ./code/razoes_social.txt relativo ao módulo.
_DEFAULT_BIBLIOTECA_PATH = Path(__file__).parent / "code" / "razoes_social.txt"
BIBLIOTECA_RAZOES_PATH = Path(
    os.environ.get("ICMS_API_BIBLIOTECA_RAZOES", str(_DEFAULT_BIBLIOTECA_PATH))
)


def carregar_biblioteca_razoes() -> dict[str, str]:
    """Carrega a biblioteca de substituição de razões sociais.

    Arquivo texto, UTF-8, uma entrada por linha no formato:
        RAZAO ORIGINAL UPPERCASE => Razao Substituta
    Linhas sem "=>" ou em branco são ignoradas.
    """
    biblioteca: dict[str, str] = {}
    if not BIBLIOTECA_RAZOES_PATH.exists():
        logger.info("Biblioteca de razões não encontrada em %s (opcional).", BIBLIOTECA_RAZOES_PATH)
        return biblioteca
    try:
        with open(BIBLIOTECA_RAZOES_PATH, encoding="utf-8") as f:
            for linha in f:
                if "=>" not in linha:
                    continue
                original, substituto = linha.strip().split("=>", maxsplit=1)
                biblioteca[original.strip().upper()] = substituto.strip()
        logger.info("Biblioteca de razões carregada: %d entradas.", len(biblioteca))
    except Exception:
        logger.exception("Falha ao carregar biblioteca de razões.")
    return biblioteca

# Parser lxml seguro contra entidades externas/explosivas.
_XML_PARSER = etree.XMLParser(
    resolve_entities=False,
    no_network=True,
    huge_tree=False,
    load_dtd=False,
)


# ===================== Auth simples =====================

def _check_api_key() -> Response | None:
    """Valida X-API-Key se ICMS_API_KEY estiver configurada."""
    if not API_KEY:
        return None
    sent = request.headers.get("X-API-Key", "").strip()
    if sent != API_KEY:
        logger.warning("Tentativa de acesso sem X-API-Key válida de %s", request.remote_addr)
        return jsonify({"error": "Unauthorized"}), 401
    return None


# ===================== Núcleo de processamento =====================

def extrair_dados_filtrados(
    xml_path: Path,
) -> tuple[str, str, dict[str, list[list[str]]]]:
    """Extrai metadados e linhas filtradas de um XML de NFe.

    Retorna (razao_social, periodo_mm_yyyy, resultados_por_grupo).
    """
    try:
        tree = etree.parse(str(xml_path), _XML_PARSER)
        root = tree.getroot()

        inf_nfe = root.find(".//nfe:infNFe", namespaces=NS)
        if inf_nfe is None:
            return "", "", {}

        ide = inf_nfe.find("nfe:ide", namespaces=NS)
        emit = inf_nfe.find("nfe:emit", namespaces=NS)
        dest = inf_nfe.find("nfe:dest", namespaces=NS)

        fornecedor = emit.findtext("nfe:xFant", default="", namespaces=NS)
        razao_social = dest.findtext("nfe:xNome", default="", namespaces=NS)
        razao_social = unescape(razao_social).replace("&", "&")
        dh_emi = ide.findtext("nfe:dhEmi", namespaces=NS)
        data_emi = datetime.strptime(dh_emi[:10], "%Y-%m-%d")
        periodo = data_emi.strftime("%m-%Y")

        uf = ide.findtext("nfe:cUF", namespaces=NS)
        chave = root.findtext(".//nfe:infProt/nfe:chNFe", namespaces=NS)
        numero_nf = ide.findtext("nfe:cNF", namespaces=NS)

        resultados: dict[str, list[list[str]]] = {nome: [] for nome in GRUPOS}

        for det in inf_nfe.findall("nfe:det", namespaces=NS):
            prod = det.find("nfe:prod", namespaces=NS)
            imposto = det.find("nfe:imposto", namespaces=NS)

            xprod = prod.findtext("nfe:xProd", default="", namespaces=NS)
            cfop = prod.findtext("nfe:CFOP", default="", namespaces=NS)
            vprod = prod.findtext("nfe:vProd", default="0", namespaces=NS)
            ncm = prod.findtext("nfe:NCM", default="", namespaces=NS)

            total = inf_nfe.find("nfe:total", namespaces=NS)
            icms_tot = total.find("nfe:ICMSTot", namespaces=NS) if total is not None else None

            v_frete = icms_tot.findtext("nfe:vFrete", namespaces=NS) if icms_tot is not None else ""
            v_outro = icms_tot.findtext("nfe:vOutro", namespaces=NS) if icms_tot is not None else ""
            v_ipi = imposto.findtext(".//nfe:IPI/nfe:IPITrib/nfe:vIPI", namespaces=NS)

            v_frete = "0" if not v_frete or float(v_frete) == 0 else v_frete
            v_outro = "0" if not v_outro or float(v_outro) == 0 else v_outro
            v_ipi = "0" if not v_ipi or float(v_ipi) == 0 else v_ipi

            cst = ""
            csosn = ""
            for icms in imposto.findall(".//nfe:ICMS", namespaces=NS):
                for child in icms.iterchildren():
                    cst = child.findtext("nfe:CST", namespaces=NS) or ""
                    csosn = child.findtext("nfe:CSOSN", namespaces=NS) or ""
                    if cst or csosn:
                        break
                if cst or csosn:
                    break

            if uf != UF_VALIDO or cfop not in CFOP_VALIDOS:
                continue

            linha = [chave, uf, numero_nf, fornecedor, xprod, ncm, cfop, cst or csosn,
                     v_frete, v_outro, v_ipi, vprod]

            for nome_saida, filtros in GRUPOS.items():
                if (filtros["cst"] and cst in filtros["cst"]) or (
                    filtros["csosn"] and csosn in filtros["csosn"]
                ):
                    resultados[nome_saida].append(linha)

        return razao_social, periodo, resultados

    except etree.XMLSyntaxError as exc:
        logger.warning("XML inválido em %s: %s", xml_path.name, exc)
        return "", "", {}
    except Exception:  # pragma: no cover — guarda-chuva defensivo
        logger.exception("Erro inesperado processando %s", xml_path.name)
        return "", "", {}


def escrever_dados_na_planilha(aba, dados: list[list[str]], celula_inicial: str) -> None:
    """Escreve linhas na aba a partir da célula inicial preservando merged cells."""
    col_letra, lin_base = coordinate_from_string(celula_inicial)
    col_index = column_index_from_string(col_letra)

    for i, linha in enumerate(dados):
        for j, valor in enumerate(linha):
            row = lin_base + i
            col = col_index + j
            cell = aba.cell(row=row, column=col)

            if isinstance(cell, MergedCell):
                continue

            if j in (8, 9, 10, 11):
                try:
                    cell.value = float(str(valor).replace(",", "."))
                    cell.number_format = 'R$ #,##0.00'
                except (ValueError, AttributeError):
                    cell.value = valor
            else:
                cell.value = valor


def _normalizar_razao(nome: str) -> str:
    nome = nome.upper().strip()
    nome = re.sub(r"[-–—]\s*ME$", "", nome)
    nome = re.sub(r"\s{2,}", " ", nome)
    return nome


def _sanitizar_nome_arquivo(nome: str) -> str:
    """Remove caracteres inválidos do Windows e força ASCII seguro."""
    nome = re.sub(r"[<>:\"/\\|?*\x00-\x1f]", "", nome)
    nome = nome.strip(" .")  # Windows não aceita trailing space/dot
    return nome or "ICMS_ST"


# ===================== Endpoints =====================

@app.errorhandler(RequestEntityTooLarge)
def _too_large(_exc):
    return jsonify({"error": f"Payload excede {MAX_UPLOAD_MB} MB"}), 413


@app.before_request
def _auth_gate():
    # Health não exige API key — útil para liveness probe.
    if request.endpoint in {"health"}:
        return None
    return _check_api_key()


@app.route("/api/icms/process", methods=["POST"])
def process_icms():
    """Recebe modelo + XMLs e devolve planilha gerada."""
    TEMP_DIR.mkdir(exist_ok=True)
    session_dir = TEMP_DIR / f"session_{datetime.now().strftime('%Y%m%d_%H%M%S_%f')}"
    session_dir.mkdir(exist_ok=True)

    try:
        if "modelo" not in request.files:
            return jsonify({"error": "Arquivo modelo não fornecido"}), 400

        modelo_file = request.files["modelo"]
        xml_files = request.files.getlist("xmls")

        if not xml_files:
            return jsonify({"error": "Nenhum arquivo XML fornecido"}), 400
        if len(xml_files) > MAX_XML_FILES:
            return jsonify({
                "error": f"Limite de {MAX_XML_FILES} XMLs por requisição excedido.",
            }), 400

        modelo_filename = secure_filename(modelo_file.filename or "")
        if not modelo_filename.lower().endswith(".xlsx"):
            return jsonify({"error": "O arquivo modelo deve ser um .xlsx"}), 400

        modelo_path = session_dir / modelo_filename
        modelo_file.save(str(modelo_path))

        xml_paths: list[Path] = []
        for xml_file in xml_files:
            xml_filename = secure_filename(xml_file.filename or "")
            if not xml_filename.lower().endswith(".xml"):
                return jsonify({
                    "error": f'Arquivo inválido: "{xml_filename}". Apenas .xml.',
                }), 400
            xml_path = session_dir / xml_filename
            xml_file.save(str(xml_path))
            xml_paths.append(xml_path)

        resultados_finais: dict[str, list[list[str]]] = {nome: [] for nome in GRUPOS}
        razoes_sociais: list[str] = []
        periodo = ""

        for xml_path in xml_paths:
            razao, data, resultados = extrair_dados_filtrados(xml_path)
            if razao:
                razoes_sociais.append(razao.strip())
            periodo = data or periodo
            for nome, linhas in resultados.items():
                resultados_finais[nome].extend(linhas)

        if not razoes_sociais or not periodo:
            return jsonify({
                "error": "Não foi possível extrair razão social/data de apuração dos XMLs.",
            }), 400

        # Aplica biblioteca de razões (substituição manual) ANTES da contagem.
        biblioteca = carregar_biblioteca_razoes()
        razoes_substituidas = []
        for razao in razoes_sociais:
            normalizado = _normalizar_razao(razao)
            substituto = biblioteca.get(normalizado, razao.strip())
            razoes_substituidas.append(substituto)
        razao_social = Counter(razoes_substituidas).most_common(1)[0][0]

        wb = load_workbook(str(modelo_path))
        if "ICMS ST 1104" not in wb.sheetnames:
            return jsonify({
                "error": f"Aba 'ICMS ST 1104' não encontrada. Abas: {', '.join(wb.sheetnames)}",
            }), 400

        aba_icms = wb["ICMS ST 1104"]
        aba_icms["C3"] = razao_social
        aba_icms["C5"] = datetime.strptime(periodo, "%m-%Y")
        aba_icms["C5"].number_format = "mmm-yy"

        for nome_grupo, (nome_aba, celula) in MAPEAMENTO_ABAS_CELULAS.items():
            dados = resultados_finais[nome_grupo]
            if not dados or nome_aba not in wb.sheetnames:
                continue
            escrever_dados_na_planilha(wb[nome_aba], dados, celula)

        razao_sanitizada = _sanitizar_nome_arquivo(razao_social)
        nome_planilha = secure_filename(f"ICMS ST {periodo}_{razao_sanitizada}.xlsx")
        arquivo_saida = session_dir / nome_planilha
        wb.save(str(arquivo_saida))

        logger.info("Processados %d XMLs para %s/%s", len(xml_paths), razao_social, periodo)

        return send_file(
            str(arquivo_saida),
            as_attachment=True,
            download_name=nome_planilha,
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )

    except Exception:
        logger.exception("Falha ao processar requisição ICMS")
        return jsonify({"error": "Erro interno ao processar XMLs"}), 500

    finally:
        # Sempre remove o diretório da sessão. send_file já carregou bytes em memória/stream.
        try:
            shutil.rmtree(session_dir, ignore_errors=True)
        except Exception:
            logger.exception("Falha ao limpar %s", session_dir)


@app.route("/api/icms/health", methods=["GET"])
def health():
    return jsonify({"status": "ok", "service": "ICMS ST API"})


@app.route("/api/icms/biblioteca", methods=["GET"])
def biblioteca_get():
    """Retorna a biblioteca de razões atual (texto bruto + entries parseadas)."""
    if not BIBLIOTECA_RAZOES_PATH.exists():
        return jsonify({"path": str(BIBLIOTECA_RAZOES_PATH), "raw": "", "entries": {}})
    raw = BIBLIOTECA_RAZOES_PATH.read_text(encoding="utf-8")
    return jsonify({
        "path": str(BIBLIOTECA_RAZOES_PATH),
        "raw": raw,
        "entries": carregar_biblioteca_razoes(),
    })


@app.route("/api/icms/biblioteca", methods=["POST"])
def biblioteca_post():
    """Sobrescreve a biblioteca de razões (admin)."""
    payload = request.get_json(silent=True) or {}
    raw = payload.get("raw", "")
    if not isinstance(raw, str):
        return jsonify({"error": "Campo 'raw' deve ser string."}), 400
    BIBLIOTECA_RAZOES_PATH.parent.mkdir(parents=True, exist_ok=True)
    BIBLIOTECA_RAZOES_PATH.write_text(raw, encoding="utf-8")
    entries = carregar_biblioteca_razoes()
    return jsonify({"saved": True, "entries_count": len(entries)})


# ===================== Entrypoint =====================

if __name__ == "__main__":
    # Bind padrão em 127.0.0.1: expor na rede exige passar ICMS_API_HOST=0.0.0.0 explicitamente.
    host = os.environ.get("ICMS_API_HOST", "127.0.0.1")
    port = int(os.environ.get("ICMS_API_PORT", "5000"))
    debug = os.environ.get("ICMS_API_DEBUG", "false").lower() == "true"

    logger.info("API ICMS ST iniciando em %s:%d (debug=%s)", host, port, debug)
    logger.info("CORS origins: %s", ALLOWED_ORIGINS)
    logger.info("API key obrigatória: %s", "sim" if API_KEY else "não")
    logger.info("Limite de upload: %d MB", MAX_UPLOAD_MB)
    logger.info("Diretório temporário: %s", TEMP_DIR)

    # Em produção usar `waitress-serve --host=127.0.0.1 --port=5000 api_icms:app`
    app.run(debug=debug, host=host, port=port)
